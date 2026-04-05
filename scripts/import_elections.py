#!/usr/bin/env python3
"""
Választási eredmények letöltése és importálása a valasztas.hu-ról.

Adatforrások:
- 2022: VTR JSON API (vtr.valasztas.hu/ogy2022/data/)
- 2018: Beágyazva a 2022-es VTR-ben (ElozoOevkEredmenyek.json)
- 2014/2010/2006: Master ZIP (static.valasztas.hu)

Futtatás:
    cd scripts && pip install -r requirements.txt && python import_elections.py
"""

import json
import zipfile
import csv
import io
from pathlib import Path
from common import get_db, log, normalize_party_name, download_file, ensure_download_dir

# === VTR 2022 JSON API URLs ===
VTR_2022_BASE = "https://vtr.valasztas.hu/ogy2022/data"
VTR_2022_CONFIG = f"{VTR_2022_BASE}/config.json"
VTR_2022_PARTIES = f"{VTR_2022_BASE}/04022333/ver/Szervezetek.json"
VTR_2022_CANDIDATES = f"{VTR_2022_BASE}/04022333/ver/EgyeniJeloltek.json"
VTR_2022_OEVK_RESULTS = f"{VTR_2022_BASE}/04161400/szavossz/OevkJkv.json"
VTR_2022_LIST_RESULTS = f"{VTR_2022_BASE}/04161400/szavossz/ListasJkv.json"
VTR_2022_PARTY_TOTALS = f"{VTR_2022_BASE}/04161400/szavossz/SzervezetekEredmenye.json"
VTR_2022_OEVK_DEFS = f"{VTR_2022_BASE}/04022333/ver/OevkAdatok.json"
VTR_2022_PREV_RESULTS = f"{VTR_2022_BASE}/04022333/ver/ElozoOevkEredmenyek.json"
VTR_2022_COUNTIES = f"{VTR_2022_BASE}/04022333/ver/Megyek.json"

# Master archive for older elections
MASTER_ZIP_URL = "https://static.valasztas.hu/dyn/letoltesek/valasztasi_eredmenyek_1990-2024.zip"


def download_vtr_json(url: str, filename: str) -> list:
    """VTR JSON fájl letöltése és parse-olása. A VTR { "header": ..., "list": [...] } formátumot használ."""
    path = download_file(url, filename)
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict) and "list" in data:
        return data["list"]
    return data if isinstance(data, list) else [data]


def build_party_map_2022(parties_data: list) -> dict[str, str]:
    """2022 VTR szervezet kódok → parties.id leképezés."""
    party_map = {}
    for p in parties_data:
        szkod = str(p.get("szkod", ""))
        name = p.get("nev", "") or p.get("r_nev", "")
        party_id = normalize_party_name(name)
        party_map[szkod] = party_id
        # Rövid név is
        short = p.get("r_nev", "")
        if short:
            party_map[short.lower()] = party_id
    return party_map


def build_county_map(counties_data: list) -> dict[str, str]:
    """Megye kódok → megye nevek."""
    county_map = {}
    for c in counties_data:
        maz = str(c.get("maz", "")).zfill(2)
        county_map[maz] = c.get("megye_nev", f"Megye {maz}")
    return county_map


def import_2022(conn):
    """2022-es választási eredmények importálása VTR JSON API-ból."""
    log.info("=== 2022 választás importálása ===")

    # Letöltések
    parties_data = download_vtr_json(VTR_2022_PARTIES, "vtr2022_szervezetek.json")
    candidates_data = download_vtr_json(VTR_2022_CANDIDATES, "vtr2022_egyeni_jeloltek.json")
    oevk_results = download_vtr_json(VTR_2022_OEVK_RESULTS, "vtr2022_oevk_jkv.json")
    list_results = download_vtr_json(VTR_2022_LIST_RESULTS, "vtr2022_listas_jkv.json")
    oevk_defs = download_vtr_json(VTR_2022_OEVK_DEFS, "vtr2022_oevk_adatok.json")
    counties_data = download_vtr_json(VTR_2022_COUNTIES, "vtr2022_megyek.json")
    party_totals = download_vtr_json(VTR_2022_PARTY_TOTALS, "vtr2022_szervezetek_eredmenye.json")

    party_map = build_party_map_2022(parties_data)
    county_map = build_county_map(counties_data)

    # Jelölt → párt leképezés (jlcs_kod = jelölőcsoport kód, jelolo_szervezetek = szervezet kódok listája)
    candidate_party = {}
    for c in candidates_data:
        ej_id = str(c.get("ej_id", ""))
        # A jelölő szervezetek kódok listája
        org_codes = c.get("jelolo_szervezetek", [])
        jlcs_nev = c.get("jlcs_nev", "")
        # Az első szervezet kódja alapján határozzuk meg a pártot
        # Ha a jelölőcsoport neve ismert, azt használjuk
        party_id = "other"
        if jlcs_nev:
            party_id = normalize_party_name(jlcs_nev)
        if party_id == "other" and org_codes:
            for code in (org_codes if isinstance(org_codes, list) else [org_codes]):
                pid = party_map.get(str(code), "other")
                if pid != "other":
                    party_id = pid
                    break
        candidate_party[ej_id] = {
            "party_id": party_id,
            "name": c.get("neve", "") or c.get("nev", ""),
            "independent": str(c.get("szkod", "")) == "0",
        }

    # Korábbi OEVK eredmények törlése (ha vannak)
    conn.execute("DELETE FROM oevk_results WHERE election_year = 2022")
    conn.execute("DELETE FROM list_results WHERE election_year = 2022")

    # OEVK eredmények importálása (egyeni_jkv.tetelek[])
    oevk_count = 0
    for oevk in oevk_results:
        maz = str(oevk.get("maz", "")).zfill(2)
        evk = str(oevk.get("evk", "")).zfill(2)
        oevk_id = f"{maz}_{evk}"

        egyeni_jkv = oevk.get("egyeni_jkv", {})
        tetelek = egyeni_jkv.get("tetelek", [])
        if not tetelek:
            continue

        # Érvényes szavazatok összesen
        total_valid = egyeni_jkv.get("szl_ervenyes", 0) or sum(t.get("szavazat", 0) for t in tetelek)

        # Rendezés szavazatszám szerint csökkenő
        tetelek_sorted = sorted(tetelek, key=lambda t: t.get("szavazat", 0), reverse=True)

        for i, t in enumerate(tetelek_sorted):
            ej_id = str(t.get("ej_id", ""))
            votes = t.get("szavazat", 0)
            cand_info = candidate_party.get(ej_id, {"party_id": "other", "name": "", "independent": False})
            vote_pct = (votes / total_valid * 100) if total_valid > 0 else 0

            conn.execute(
                """INSERT INTO oevk_results
                   (election_year, oevk_id, oevk_id_2026, party_id, candidate_name, votes, vote_share_pct, is_winner)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (2022, oevk_id, oevk_id, cand_info["party_id"],
                 cand_info["name"], votes, round(vote_pct, 2), 1 if i == 0 else 0)
            )
            oevk_count += 1

    log.info(f"  {oevk_count} OEVK jelölt eredmény importálva")

    # Listás eredmények importálása a SzervezetekEredmenye-ből
    # FONTOS: Koalíciós pártok ugyanazt a tl_id-t és szavazatszámot kapják.
    # Deduplikálás tl_id alapján — egy lista csak egyszer számít.

    # Koalíciós listák azonosítása: az egyéni jelöltek jelölőcsoport-kódjaiból
    # Azok a listák amelyek több pártot jelölnek, koalíciós listák.
    # tl_id=952: Fidesz-KDNP közös lista → fidesz_kdnp
    # tl_id=950: DK-JOBBIK-MOMENTUM-MSZP-LMP-PÁRBESZÉD lista → egyseges_ellenzek
    list_party_overrides: dict[int, str] = {}
    for c in candidates_data:
        jlcs_nev = c.get("jlcs_nev", "")
        # Ha a jelöltnek van listája, a lista tl_id-ját keressük
        listak = c.get("listak", [])
        if not jlcs_nev or not listak:
            continue
        party_id = normalize_party_name(jlcs_nev)
        if party_id in ("other",):
            continue
        for lista in listak:
            if isinstance(lista, dict):
                tid = lista.get("tl_id")
            elif isinstance(lista, (int, str)):
                tid = int(lista)
            else:
                continue
            if tid and tid not in list_party_overrides:
                list_party_overrides[tid] = party_id

    list_count = 0
    seen_tl_ids: set[int] = set()
    unique_list_entries: list[tuple[str, int]] = []  # (party_id, votes)

    for p in party_totals:
        szkod = str(p.get("szkod", ""))
        votes = p.get("listas_szavazat", 0)
        tl_id = p.get("tl_id")
        if votes <= 0:
            continue
        # Független jelöltek (szkod=0) nem indulnak listán
        if szkod == "0":
            continue
        # Deduplikálás tl_id alapján
        if tl_id and tl_id in seen_tl_ids:
            continue
        if tl_id:
            seen_tl_ids.add(tl_id)

        # Party ID: override ha koalíciós lista, egyébként szervezet kódból
        party_id = list_party_overrides.get(tl_id, party_map.get(szkod, "other"))
        unique_list_entries.append((party_id, votes))

    total_list_votes = sum(v for _, v in unique_list_entries)
    for party_id, votes in unique_list_entries:
        vote_pct = (votes / total_list_votes * 100) if total_list_votes > 0 else 0
        conn.execute(
            """INSERT INTO list_results
               (election_year, level, county, party_id, votes, vote_share_pct)
               VALUES (?, 'national', NULL, ?, ?, ?)""",
            (2022, party_id, votes, round(vote_pct, 2))
        )
        list_count += 1

    log.info(f"  {list_count} listás eredmény importálva (deduplikálva tl_id alapján)")
    conn.commit()


def import_2018_from_vtr(conn):
    """2018-as OEVK eredmények importálása a 2022-es VTR ElozoOevkEredmenyek.json-ból."""
    log.info("=== 2018 választás importálása (VTR 2022-ből) ===")

    prev_data = download_vtr_json(VTR_2022_PREV_RESULTS, "vtr2022_elozo_oevk_eredmenyek.json")
    counties_data = download_vtr_json(VTR_2022_COUNTIES, "vtr2022_megyek.json")
    county_map = build_county_map(counties_data)

    conn.execute("DELETE FROM oevk_results WHERE election_year = 2018")

    # Csoportosítás OEVK-nként
    oevk_groups: dict[str, list] = {}
    for entry in prev_data:
        maz = str(entry.get("maz", "")).zfill(2)
        evk = entry.get("evk", 0)
        oevk_id = f"{maz}_{str(evk).zfill(2)}"
        if oevk_id not in oevk_groups:
            oevk_groups[oevk_id] = []
        oevk_groups[oevk_id].append(entry)

    count = 0
    for oevk_id, candidates in oevk_groups.items():
        # Rendezés szavazatszám szerint
        candidates_sorted = sorted(candidates, key=lambda c: c.get("szavazat", 0), reverse=True)
        total = sum(c.get("szavazat", 0) for c in candidates_sorted)

        for i, c in enumerate(candidates_sorted):
            name = c.get("neve", "") or c.get("nev", "")
            party_name = c.get("jlcs_nev", "") or c.get("szervezet_nev", "") or c.get("jelolo_szervezetek", "")
            votes = c.get("szavazat", 0)
            vote_pct = (votes / total * 100) if total > 0 else 0
            party_id = normalize_party_name(party_name)

            conn.execute(
                """INSERT INTO oevk_results
                   (election_year, oevk_id, oevk_id_2026, party_id, candidate_name, votes, vote_share_pct, is_winner)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (2018, oevk_id, oevk_id, party_id, name, votes, round(vote_pct, 2), 1 if i == 0 else 0)
            )
            count += 1

    log.info(f"  {count} OEVK jelölt eredmény importálva (2018)")
    conn.commit()


def import_from_master_zip(conn):
    """
    Régebbi választások importálása a master ZIP-ből.
    A ZIP nested: outer.zip → YYYY_parlamenti.zip → XLSX fájlok
    """
    import openpyxl

    log.info("=== Régebbi választások importálása (master ZIP) ===")

    zip_path = ensure_download_dir() / "valasztasi_eredmenyek_1990-2024.zip"
    if not zip_path.exists():
        log.info("Master ZIP letöltése (295 MB, ez eltarthat pár percig)...")
        download_file(MASTER_ZIP_URL, "valasztasi_eredmenyek_1990-2024.zip")

    if not zip_path.exists():
        log.warning("Master ZIP nem elérhető, kihagyjuk a régebbi választásokat.")
        return

    years_to_import = {
        2006: "2006_parlamenti.zip",
        2010: "2010_parlamenti.zip",
        2014: "2014_parlamenti.zip",
    }

    try:
        with zipfile.ZipFile(zip_path, "r") as outer:
            for year, inner_name in years_to_import.items():
                full_name = f"valasztasi_eredmenyek_1990-2024/{inner_name}"
                if full_name not in outer.namelist():
                    log.warning(f"  {year}: {inner_name} nem található a ZIP-ben")
                    continue

                log.info(f"  {year}: {inner_name} kicsomagolása...")
                inner_data = outer.read(full_name)

                with zipfile.ZipFile(io.BytesIO(inner_data)) as inner:
                    inner_files = inner.namelist()
                    log.info(f"    {len(inner_files)} fájl: {inner_files}")

                    for fname in inner_files:
                        lower = fname.lower()
                        if lower.endswith(".xlsx") and ("list" in lower or "orsz" in lower):
                            log.info(f"    Feldolgozás: {fname}")
                            xlsx_data = inner.read(fname)
                            count = _import_xlsx_list(conn, year, xlsx_data, fname)
                            if count > 0:
                                log.info(f"      {count} listás eredmény importálva")

    except zipfile.BadZipFile:
        log.error("Érvénytelen ZIP fájl!")
    except Exception as e:
        log.error(f"ZIP feldolgozási hiba: {e}")


def _import_xlsx_list(conn, year: int, xlsx_data: bytes, filename: str) -> int:
    """
    XLSX listás eredmények importálása.
    A fájl szavazóköri szintű lehet — aggregáljuk pártonként nemzeti szintre.
    """
    import openpyxl

    conn.execute("DELETE FROM list_results WHERE election_year = ?", (year,))

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=True, data_only=True)

    # Fejléc keresés az első sheet-ben
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        wb.close()
        return 0

    # Oszlopok azonosítása
    name_col = None
    votes_col = None
    for i, h in enumerate(header or []):
        if h is None:
            continue
        lower = str(h).lower()
        if name_col is None and ("szervezet" in lower or "jelölő" in lower or "párt" in lower or "lista" in lower or "nev" in lower):
            name_col = i
        elif votes_col is None and "szavazat" in lower and "érvény" not in lower:
            votes_col = i

    if name_col is None or votes_col is None:
        log.warning(f"    Nem találtam párt/szavazat oszlopot a fejlécben: {header}")
        wb.close()
        return 0

    log.info(f"    Párt oszlop: {name_col} ({header[name_col]}), Szavazat oszlop: {votes_col} ({header[votes_col]})")

    # Szavazóköri szintű adatok aggregálása pártonként
    party_totals: dict[str, int] = {}
    row_count = 0

    for row in rows_iter:
        if not row or len(row) <= max(name_col, votes_col):
            continue
        party_name = str(row[name_col] or "").strip()
        if not party_name or party_name.lower() in ("összesen", "összes", "total", ""):
            continue
        try:
            votes = int(float(str(row[votes_col] or 0).replace(" ", "").replace(",", "")))
        except (ValueError, TypeError):
            continue
        if votes <= 0:
            continue

        party_id = normalize_party_name(party_name)
        party_totals[party_id] = party_totals.get(party_id, 0) + votes
        row_count += 1

    wb.close()
    log.info(f"    {row_count} szavazóköri sor aggregálva → {len(party_totals)} párt")

    # Aggregált eredmények beszúrása
    total_votes = sum(party_totals.values())
    count = 0
    for party_id, votes in sorted(party_totals.items(), key=lambda x: -x[1]):
        pct = (votes / total_votes * 100) if total_votes > 0 else 0
        conn.execute(
            """INSERT INTO list_results
               (election_year, level, county, party_id, votes, vote_share_pct)
               VALUES (?, 'national', NULL, ?, ?, ?)""",
            (year, party_id, votes, round(pct, 2))
        )
        count += 1

    if count > 0:
        conn.commit()
    return count


def main():
    conn = get_db()
    try:
        # 2022 — VTR JSON API (legjobb minőségű adat)
        import_2022(conn)

        # 2018 — beágyazva a 2022-es VTR-ben
        import_2018_from_vtr(conn)

        # 2014, 2010, 2006 — master ZIP (ha elérhető)
        import_from_master_zip(conn)

        # Végső statisztika
        for year in [2006, 2010, 2014, 2018, 2022]:
            oevk_count = conn.execute(
                "SELECT COUNT(*) FROM oevk_results WHERE election_year = ?", (year,)
            ).fetchone()[0]
            list_count = conn.execute(
                "SELECT COUNT(*) FROM list_results WHERE election_year = ?", (year,)
            ).fetchone()[0]
            log.info(f"  {year}: {oevk_count} OEVK + {list_count} listás eredmény")

    finally:
        conn.close()

    log.info("Kész!")


if __name__ == "__main__":
    main()
